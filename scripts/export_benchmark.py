#!/usr/bin/env python3
"""
Pull benchmark CSV from Android device and generate a polished Excel report.

Usage:
    python3 scripts/export_benchmark.py [output.xlsx]
    python3 scripts/export_benchmark.py --csv /path/to/local.csv [output.xlsx]

Requirements:
    pip3 install openpyxl
"""
import subprocess
import csv
import sys
import os
import re
import math
from collections import OrderedDict
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_from_string
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip3 install openpyxl")
    sys.exit(1)


# -----------------------------------------------------------------------
# Styling constants
# -----------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HALIDE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
OPENCV_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
SPEEDUP_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
WINNER_HALIDE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WINNER_OPENCV_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
CATEGORY_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

# Heatmap gradient fills for speedup values
HEAT_DARK_GREEN = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
HEAT_MED_GREEN = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
HEAT_LIGHT_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
HEAT_LIGHT_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEAT_DARK_RED = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=12, bold=True, color="2E75B6")
META_LABEL_FONT = Font(name="Calibri", size=11, bold=True, color="333333")
META_VALUE_FONT = Font(name="Calibri", size=11, color="333333")
CATEGORY_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
NUM_FMT = "#,##0"
PCT_FMT = "0.0%"

# Operation categories for heatmap grouping
CATEGORIES = OrderedDict([
    ("Color Conversion", ["RGB to BGR", "NV21 to RGB", "RGB to NV21"]),
    ("Blur", ["Gaussian Blur (5x5)", "Lens Blur (r=4)"]),
    ("Resize (Scale)", ["Resize Bilinear (0.5x)", "Resize Bicubic (0.5x)", "Resize Area (0.5x)",
                        "Resize Letterbox (720p)"]),
    ("Resize (Target)", ["Resize Bilinear Target (720p)", "Resize Bicubic Target (720p)",
                         "Resize Area Target (720p)"]),
    ("Rotate", ["Rotate 90", "Rotate Arbitrary (45\u00b0)"]),
    ("Flip", ["Flip Horizontal", "Flip Vertical"]),
    ("Fused Pipeline", ["NV21 Pipeline Bilinear (rotate+resize)",
                        "NV21 Pipeline Area (rotate+resize)"]),
])


def pull_csv_from_device():
    """Pull CSV from Android device via adb."""
    local_path = "/tmp/benchmark_results.csv"
    result = subprocess.run(
        [
            "adb", "pull",
            "/sdcard/Android/data/com.example.halidetest/files/benchmark_results.csv",
            local_path,
        ],
        capture_output=True, text=True,
    )
    if result.returncode != 0:
        print(f"adb pull failed: {result.stderr.strip()}")
        print("Make sure a device is connected and the app has run benchmarks.")
        sys.exit(1)
    print(f"Pulled CSV to {local_path}")
    return local_path


def get_device_info():
    """Query device model and manufacturer via adb. Returns dict."""
    info = {"model": "Unknown", "manufacturer": "Unknown"}
    for prop, key in [("ro.product.model", "model"),
                      ("ro.product.manufacturer", "manufacturer")]:
        try:
            result = subprocess.run(
                ["adb", "shell", "getprop", prop],
                capture_output=True, text=True, timeout=5,
            )
            if result.returncode == 0 and result.stdout.strip():
                info[key] = result.stdout.strip()
        except (subprocess.TimeoutExpired, FileNotFoundError):
            pass
    return info


def read_csv(csv_path):
    """Read CSV into list of dicts. Auto-detect header or add one."""
    rows = []
    with open(csv_path, newline="") as f:
        sample = f.read(512)
        f.seek(0)
        if sample.startswith("operation"):
            reader = csv.DictReader(f)
        else:
            reader = csv.DictReader(
                f,
                fieldnames=[
                    "operation", "framework", "resolution",
                    "median_us", "mean_us", "min_us", "max_us", "timestamp",
                ],
            )
        for row in reader:
            rows.append(row)
    return rows


def compute_summary(rows):
    """Group rows by (operation, resolution) and compute summary metrics."""
    data = OrderedDict()
    for row in rows:
        key = (row.get("operation", ""), row.get("resolution", ""))
        fw = row.get("framework", "")
        if key not in data:
            data[key] = {}
        data[key][fw] = row

    summary = []
    for (op, res), frameworks in data.items():
        h = frameworks.get("Halide", {})
        o = frameworks.get("OpenCV", {})
        h_med = int(h.get("median_us", 0))
        o_med = int(o.get("median_us", 0))
        h_mean = int(h.get("mean_us", 0))
        o_mean = int(o.get("mean_us", 0))
        speedup = o_med / h_med if h_med > 0 else 0

        if h_med > 0 and o_med > 0:
            diff_us = o_med - h_med  # positive = Halide faster
            winner = "Halide" if diff_us > 0 else "OpenCV"
            faster = min(h_med, o_med)
            slower = max(h_med, o_med)
            diff_pct = (slower - faster) / slower if slower > 0 else 0
        else:
            diff_us = 0
            winner = "N/A"
            diff_pct = 0

        summary.append({
            "op": op, "res": res,
            "h_med": h_med, "o_med": o_med,
            "h_mean": h_mean, "o_mean": o_mean,
            "speedup": speedup,
            "diff_us": diff_us, "diff_pct": diff_pct,
            "winner": winner,
        })
    return summary


def style_header_row(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)


def speedup_fill(val):
    """Return a fill color based on speedup value."""
    if val >= 2.0:
        return HEAT_DARK_GREEN
    elif val >= 1.5:
        return HEAT_MED_GREEN
    elif val >= 1.0:
        return HEAT_LIGHT_GREEN
    elif val >= 0.75:
        return HEAT_LIGHT_RED
    else:
        return HEAT_DARK_RED


def speedup_font(val):
    """Return font for speedup cells (white text on dark backgrounds)."""
    if val >= 2.0 or val < 0.75:
        return Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    else:
        return Font(name="Calibri", size=11, bold=True)


# -----------------------------------------------------------------------
# Sheet builders
# -----------------------------------------------------------------------

def build_dashboard(wb, summary, device_info):
    """Sheet 1: Dashboard with metadata, winners table, and overall stats."""
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = "1F4E79"

    # --- Title ---
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Halide vs OpenCV Benchmark Report"
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Metadata ---
    row = 3
    meta = [
        ("Device", device_info.get("model", "Unknown")),
        ("Manufacturer", device_info.get("manufacturer", "Unknown")),
        ("Test Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Halide Version", "21.0.0"),
        ("OpenCV Version", "3.4.16"),
        ("Total Operations", str(len(set(s["op"] for s in summary)))),
    ]
    for label, value in meta:
        ws.cell(row=row, column=1, value=label).font = META_LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = META_VALUE_FONT
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=2).alignment = LEFT
        row += 1

    # --- Overall stats ---
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Overall Performance Summary").font = SUBTITLE_FONT
    row += 1

    halide_wins = sum(1 for s in summary if s["winner"] == "Halide")
    opencv_wins = sum(1 for s in summary if s["winner"] == "OpenCV")
    speedups = [s["speedup"] for s in summary if s["speedup"] > 0]
    geo_mean = math.exp(sum(math.log(s) for s in speedups) / len(speedups)) if speedups else 0

    stats = [
        ("Halide Wins", halide_wins, WINNER_HALIDE_FILL),
        ("OpenCV Wins", opencv_wins, WINNER_OPENCV_FILL),
        ("Avg Speedup (geometric mean)", f"{geo_mean:.2f}x", SPEEDUP_FILL),
    ]
    for label, value, fill in stats:
        ws.cell(row=row, column=1, value=label).font = META_LABEL_FONT
        c = ws.cell(row=row, column=2, value=value)
        c.font = BOLD_FONT
        c.fill = fill
        c.alignment = CENTER
        c.border = THIN_BORDER
        row += 1

    # --- Winners table ---
    row += 1
    ws.merge_cells(f"A{row}:H{row}")
    ws.cell(row=row, column=1, value="Performance Comparison by Operation").font = SUBTITLE_FONT
    row += 1

    headers = ["Operation", "Resolution", "Halide (us)", "OpenCV (us)",
               "Winner", "Diff (us)", "Diff (%)", "Speedup"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER
    row += 1

    for s in summary:
        ws.cell(row=row, column=1, value=s["op"]).border = THIN_BORDER
        ws.cell(row=row, column=2, value=s["res"]).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = CENTER

        ch = ws.cell(row=row, column=3, value=s["h_med"])
        ch.number_format = NUM_FMT
        ch.fill = HALIDE_FILL
        ch.alignment = CENTER
        ch.border = THIN_BORDER

        co = ws.cell(row=row, column=4, value=s["o_med"])
        co.number_format = NUM_FMT
        co.fill = OPENCV_FILL
        co.alignment = CENTER
        co.border = THIN_BORDER

        cw = ws.cell(row=row, column=5, value=s["winner"])
        cw.fill = WINNER_HALIDE_FILL if s["winner"] == "Halide" else WINNER_OPENCV_FILL
        cw.font = BOLD_FONT
        cw.alignment = CENTER
        cw.border = THIN_BORDER

        cd = ws.cell(row=row, column=6, value=abs(s["diff_us"]))
        cd.number_format = NUM_FMT
        cd.alignment = CENTER
        cd.border = THIN_BORDER

        cp = ws.cell(row=row, column=7, value=s["diff_pct"])
        cp.number_format = PCT_FMT
        cp.alignment = CENTER
        cp.border = THIN_BORDER

        cs = ws.cell(row=row, column=8, value=f"{s['speedup']:.2f}x")
        cs.fill = SPEEDUP_FILL
        cs.alignment = CENTER
        cs.border = THIN_BORDER

        row += 1

    auto_width(ws)
    ws.column_dimensions["A"].width = 38


def build_summary(wb, summary):
    """Sheet 2: Enhanced summary pivot table."""
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "2E75B6"

    headers = ["Operation", "Resolution",
               "Halide Median (us)", "OpenCV Median (us)",
               "Diff (us)", "Diff (%)",
               "Halide Mean (us)", "OpenCV Mean (us)",
               "Speedup (median)", "Winner"]
    ws.append(headers)
    style_header_row(ws, len(headers))

    for s in summary:
        ws.append([
            s["op"], s["res"],
            s["h_med"], s["o_med"],
            abs(s["diff_us"]), s["diff_pct"],
            s["h_mean"], s["o_mean"],
            f"{s['speedup']:.2f}x", s["winner"],
        ])

    for r_idx in range(2, ws.max_row + 1):
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = THIN_BORDER
            if c_idx in (3, 5, 7):  # Halide columns + diff
                cell.number_format = NUM_FMT
                cell.alignment = CENTER
            if c_idx in (4, 8):  # OpenCV columns
                cell.number_format = NUM_FMT
                cell.alignment = CENTER
            if c_idx == 3:
                cell.fill = HALIDE_FILL
            elif c_idx == 4:
                cell.fill = OPENCV_FILL
            elif c_idx == 5:
                cell.alignment = CENTER
            elif c_idx == 6:
                cell.number_format = PCT_FMT
                cell.alignment = CENTER
            elif c_idx == 7:
                cell.fill = HALIDE_FILL
            elif c_idx == 8:
                cell.fill = OPENCV_FILL
            elif c_idx == 9:
                cell.fill = SPEEDUP_FILL
                cell.alignment = CENTER
                # Bold if speedup > 2.0
                try:
                    val = float(str(cell.value).replace("x", ""))
                    if val > 2.0:
                        cell.font = BOLD_FONT
                except (ValueError, TypeError):
                    pass
            elif c_idx == 10:
                winner = cell.value
                cell.fill = WINNER_HALIDE_FILL if winner == "Halide" else WINNER_OPENCV_FILL
                cell.font = BOLD_FONT
                cell.alignment = CENTER

    auto_width(ws)
    ws.column_dimensions["A"].width = 38


def build_heatmap(wb, summary):
    """Sheet 3: Operations grouped by category with color-coded speedup."""
    ws = wb.create_sheet("Heatmap")
    ws.sheet_properties.tabColor = "00B050"

    row = 1
    # Build lookup from summary
    lookup = {}
    for s in summary:
        lookup.setdefault(s["op"], []).append(s)

    for cat_name, ops in CATEGORIES.items():
        # Category header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cat_cell = ws.cell(row=row, column=1, value=cat_name)
        cat_cell.font = CATEGORY_FONT
        cat_cell.fill = CATEGORY_FILL
        cat_cell.alignment = LEFT
        cat_cell.border = THIN_BORDER
        for ci in range(2, 7):
            ws.cell(row=row, column=ci).fill = CATEGORY_FILL
            ws.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

        # Sub-header
        sub_headers = ["Operation", "Resolution", "Halide (us)", "OpenCV (us)",
                       "Speedup", "Winner"]
        for ci, h in enumerate(sub_headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
            c.alignment = CENTER
            c.border = THIN_BORDER
        row += 1

        # Data rows for each operation in this category
        found_any = False
        for op_name in ops:
            entries = lookup.get(op_name, [])
            for s in entries:
                found_any = True
                ws.cell(row=row, column=1, value=s["op"]).border = THIN_BORDER
                ws.cell(row=row, column=2, value=s["res"]).border = THIN_BORDER
                ws.cell(row=row, column=2).alignment = CENTER

                ch = ws.cell(row=row, column=3, value=s["h_med"])
                ch.number_format = NUM_FMT
                ch.fill = HALIDE_FILL
                ch.alignment = CENTER
                ch.border = THIN_BORDER

                co = ws.cell(row=row, column=4, value=s["o_med"])
                co.number_format = NUM_FMT
                co.fill = OPENCV_FILL
                co.alignment = CENTER
                co.border = THIN_BORDER

                sp_val = s["speedup"]
                cs = ws.cell(row=row, column=5, value=f"{sp_val:.2f}x")
                cs.fill = speedup_fill(sp_val)
                cs.font = speedup_font(sp_val)
                cs.alignment = CENTER
                cs.border = THIN_BORDER

                cw = ws.cell(row=row, column=6, value=s["winner"])
                cw.fill = WINNER_HALIDE_FILL if s["winner"] == "Halide" else WINNER_OPENCV_FILL
                cw.font = BOLD_FONT
                cw.alignment = CENTER
                cw.border = THIN_BORDER

                row += 1

        if not found_any:
            ws.cell(row=row, column=1, value="(no data)").border = THIN_BORDER
            row += 1

        # Blank separator row
        row += 1

    auto_width(ws)
    ws.column_dimensions["A"].width = 38


def build_charts(wb, summary):
    """Sheet 4: Charts with median comparison and speedup factor."""
    ws = wb.create_sheet("Charts")
    ws.sheet_properties.tabColor = "ED7D31"

    if not summary:
        ws.cell(row=1, column=1, value="No data available for charts.")
        return

    # --- Helper data area (columns J-N) for chart data ---
    ws.cell(row=1, column=10, value="Operation")
    ws.cell(row=1, column=11, value="Halide Median (us)")
    ws.cell(row=1, column=12, value="OpenCV Median (us)")
    ws.cell(row=1, column=13, value="Speedup")
    ws.cell(row=1, column=14, value="Reference (1.0x)")

    for i, s in enumerate(summary):
        r = i + 2
        ws.cell(row=r, column=10, value=s["op"])
        ws.cell(row=r, column=11, value=s["h_med"])
        ws.cell(row=r, column=12, value=s["o_med"])
        ws.cell(row=r, column=13, value=round(s["speedup"], 2))
        ws.cell(row=r, column=14, value=1.0)

    nrows = len(summary) + 1
    n = len(summary)

    # Hide helper data columns
    for col_letter in ['J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col_letter].hidden = True

    # --- Chart 1: Median comparison (horizontal grouped bar) ---
    chart1 = BarChart()
    chart1.type = "bar"
    chart1.grouping = "clustered"
    chart1.title = "Halide vs OpenCV - Median Execution Time"
    chart1.x_axis.title = "Time (microseconds)"
    chart1.x_axis.numFmt = "#,##0"
    chart1.y_axis.title = None
    chart1.y_axis.tickLblPos = "low"
    chart1.style = 10
    chart1.width = 28
    chart1.height = max(10, n * 1.5)
    chart1.gapWidth = 80
    chart1.legend.position = "b"

    halide_data = Reference(ws, min_col=11, min_row=1, max_row=nrows)
    opencv_data = Reference(ws, min_col=12, min_row=1, max_row=nrows)
    cats = Reference(ws, min_col=10, min_row=2, max_row=nrows)

    chart1.add_data(halide_data, titles_from_data=True)
    chart1.add_data(opencv_data, titles_from_data=True)
    chart1.set_categories(cats)

    chart1.series[0].graphicalProperties.solidFill = "70AD47"  # Green for Halide
    chart1.series[1].graphicalProperties.solidFill = "ED7D31"  # Orange for OpenCV

    # Data labels on bars
    for s in chart1.series:
        s.dLbls = DataLabelList()
        s.dLbls.showVal = True
        s.dLbls.numFmt = "#,##0"
        s.dLbls.showCatName = False
        s.dLbls.showSerName = False

    ws.add_chart(chart1, "A1")

    # --- Chart 2: Speedup factor (horizontal bar) with 1.0x reference ---
    from openpyxl.chart import LineChart

    chart2 = BarChart()
    chart2.type = "bar"
    chart2.title = "Speedup Factor (OpenCV / Halide)"
    chart2.x_axis.title = "Speedup (x)"
    chart2.x_axis.numFmt = "0.0"
    chart2.y_axis.title = None
    chart2.y_axis.tickLblPos = "low"
    chart2.style = 10
    chart2.width = 28
    chart2.height = max(10, n * 1.5)
    chart2.gapWidth = 100
    chart2.legend.position = "b"

    speedup_data = Reference(ws, min_col=13, min_row=1, max_row=nrows)
    speedup_cats = Reference(ws, min_col=10, min_row=2, max_row=nrows)
    chart2.add_data(speedup_data, titles_from_data=True)
    chart2.set_categories(speedup_cats)

    # Conditional coloring: green if Halide wins, red if OpenCV wins
    for i, s in enumerate(summary):
        pt = DataPoint(idx=i)
        color = "70AD47" if s["speedup"] >= 1.0 else "FF4444"
        pt.graphicalProperties.solidFill = color
        chart2.series[0].data_points.append(pt)

    # Data labels on speedup bars
    chart2.series[0].dLbls = DataLabelList()
    chart2.series[0].dLbls.showVal = True
    chart2.series[0].dLbls.numFmt = '0.00"x"'
    chart2.series[0].dLbls.showCatName = False
    chart2.series[0].dLbls.showSerName = False

    # Add 1.0x reference line
    line = LineChart()
    ref_data = Reference(ws, min_col=14, min_row=1, max_row=nrows)
    line.add_data(ref_data, titles_from_data=True)
    line.series[0].graphicalProperties.line.solidFill = "FF0000"
    line.series[0].graphicalProperties.line.width = 25000
    line.series[0].graphicalProperties.line.dashStyle = "dash"
    line.series[0].graphicalProperties.noFill = True
    chart2 += line

    chart2_anchor_row = max(18, int(chart1.height) + 2)
    ws.add_chart(chart2, f"A{chart2_anchor_row}")


def build_raw_data(wb, rows):
    """Sheet 5: Raw data (all CSV entries)."""
    ws = wb.create_sheet("Raw Data")
    ws.sheet_properties.tabColor = "A5A5A5"

    headers = ["Operation", "Framework", "Resolution",
               "Median (us)", "Mean (us)", "Min (us)", "Max (us)", "Timestamp"]
    ws.append(headers)

    for row in rows:
        ws.append([
            row.get("operation", ""),
            row.get("framework", ""),
            row.get("resolution", ""),
            int(row.get("median_us", 0)),
            int(row.get("mean_us", 0)),
            int(row.get("min_us", 0)),
            int(row.get("max_us", 0)),
            row.get("timestamp", ""),
        ])

    style_header_row(ws, len(headers))

    for r_idx in range(2, ws.max_row + 1):
        fw = ws.cell(row=r_idx, column=2).value
        fill = HALIDE_FILL if fw == "Halide" else OPENCV_FILL
        for c_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.border = THIN_BORDER
            cell.fill = fill
            if 4 <= c_idx <= 7:
                cell.number_format = NUM_FMT
                cell.alignment = CENTER

    auto_width(ws)


# -----------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------

def generate_excel(rows, output_path, device_info):
    summary = compute_summary(rows)
    wb = Workbook()

    build_dashboard(wb, summary, device_info)
    build_summary(wb, summary)
    build_heatmap(wb, summary)
    build_charts(wb, summary)
    build_raw_data(wb, rows)

    wb.save(output_path)
    print(f"Report saved to: {output_path}")


def main():
    args = sys.argv[1:]
    csv_path = None
    output_path = None
    has_device = True

    i = 0
    while i < len(args):
        if args[i] == "--csv" and i + 1 < len(args):
            csv_path = args[i + 1]
            has_device = False
            i += 2
        else:
            output_path = args[i]
            i += 1

    # Get device info (for filename and dashboard)
    if has_device:
        device_info = get_device_info()
    else:
        device_info = {"model": "Unknown", "manufacturer": "Unknown"}

    # Pull CSV if needed
    if csv_path is None:
        csv_path = pull_csv_from_device()

    # Generate dynamic filename if not specified
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        model_safe = re.sub(r"[^\w\-]", "_", device_info["model"])
        output_path = f"benchmark_{model_safe}_{timestamp}.xlsx"

    if not os.path.exists(csv_path):
        print(f"ERROR: CSV file not found: {csv_path}")
        sys.exit(1)

    rows = read_csv(csv_path)
    if not rows:
        print("ERROR: CSV file is empty or has no data rows.")
        sys.exit(1)

    print(f"Read {len(rows)} benchmark entries.")
    generate_excel(rows, output_path, device_info)


if __name__ == "__main__":
    main()
